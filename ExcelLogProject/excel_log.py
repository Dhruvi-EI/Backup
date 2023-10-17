import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import config as cf
import concurrent.futures

class ExcelGenerator:
    file_delete=[cf.excel_output]
    for i in file_delete:
        if i in cf.project_name:
            os.remove(i)
        else:
            continue

    @staticmethod
    def generate_excel():
        # Create a new workbook
        wb = Workbook()
        data_dicts = [cf.dict1, cf.dict2, cf.dict3, cf.dict4]
        ws = wb["Sheet"]
        column_names = ['Alarm remove', 'Frequency', '', 'Alarm Set', 'Frequency',
            '', 'Alarm clear', 'Frequency', '', 'Alarm confirmed', 'Frequency']
        for col, column_name in enumerate(column_names, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = Font(bold=True, size=11.5)

        # Write data
        for i, data_dict in enumerate(data_dicts):
            col_offset=i*3
            for row, (key, value) in enumerate(data_dict.items(), start=2):
                ws.cell(row=row, column=col_offset + 1, value=key)
                ws.cell(row=row, column=col_offset + 2, value=value)

        # Auto-adjust column widths to fit content
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get column letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width=(max_length+2)  # Adjust as needed
            ws.column_dimensions[column_letter].width = adjusted_width
        wb.save(cf.excel_output)
        wb.close()

    @staticmethod
    def read_file(filename):
        with open(filename, 'r') as file:
            return file.read()

    @staticmethod
    def split_digits(text1):
        l1 = text1.split("#")
        l2 = l1[1].split(' ')
        return l2[0]

    @staticmethod
    def split_confirmedByuser(text2):
        l3 = text2.split("#")
        l4 = l3[1].split(' ')
        return l4[0]

    @staticmethod
    def main():
        with open(cf.log_file, 'r') as new_log, open(cf.data_file, 'w') as a1:
            for line in new_log:
                cf.list2.append(line)
                for j in cf.list1:
                    if j in line.lower():
                        a1.write(line)

        with concurrent.futures.ThreadPoolExecutor() as executor:
            results = list(executor.map(ExcelGenerator.read_file, cf.filenames))

        for i in range(len(results)):
            results2 = results[i].split("\n")
            cf.results1.extend(results2)
        print(cf.results1)

        for i in cf.results1:
            if i == '':
                continue
            cf.list6.append(i)
            if 'event remove' in i:
                data1 = ExcelGenerator.split_digits(cf.list6[0])
                data1 = int(data1)
                if data1 in cf.dict1.keys():
                    cf.dict1[data1] += 1
                else:
                    cf.dict1[data1] = 1
            elif 'event set' in i:
                data2 = ExcelGenerator.split_digits(cf.list6[0])
                data2 = int(data2)
                if data2 in cf.dict2.keys():
                    cf.dict2[data2] += 1
                else:
                    cf.dict2[data2] = 1
            elif 'event clear' in i:
                data3 = ExcelGenerator.split_digits(cf.list6[0])
                data3 = int(data3)
                if data3 in cf.dict3.keys():
                    cf.dict3[data3] += 1
                else:
                    cf.dict3[data3] = 1
            elif 'confirmed by user' in i:
                data4 = ExcelGenerator.split_confirmedByuser(cf.list6[0])
                data4 = int(data4)
                if data4 in cf.dict4.keys():
                    cf.dict4[data4] += 1
                else:
                    cf.dict4[data4] = 1
            else:
                print("entry not found")
            cf.list6 = []
        ExcelGenerator.generate_excel() # Generate Excel

        file_to_delete=[cf.data_file]  # Specify the path to the file you want to delete
        for i in file_to_delete:
            os.remove(i)
            print(f"{i} has been deleted successfully.")
ExcelGenerator.main()